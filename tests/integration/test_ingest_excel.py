from pathlib import Path

import openpyxl
import pytest

from planer.adapters.ingest_excel import parse_excel


def test_parse_correct_participant_count(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    assert len(participants) == 3


def test_parse_participant_fields(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    alice = next(p for p in participants if p.user_id == "u1")
    assert alice.display_name == "Alice Muster"
    assert alice.email == "alice@uni.de"
    assert alice.matr_nr == "111111"
    assert alice.group_id == "g1"
    assert alice.group_name == "Gruppe 1"


def test_parse_has_admission_wahr(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    alice = next(p for p in participants if p.user_id == "u1")
    assert alice.has_admission is True


def test_parse_has_admission_falsch(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    carol = next(p for p in participants if p.user_id == "u3")
    assert carol.has_admission is False


def test_parse_result_fields(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    alice = next(p for p in participants if p.user_id == "u1")
    assert len(alice.results) == 1
    r = alice.results[0]
    assert r.rule == "INDIVIDUAL_PERCENT_WITH_ALLOWED_FAILURES"
    assert r.assignment_type == "HOMEWORK"
    assert r.achieved_points == 2
    assert r.passed is False


def test_parse_result_passed_wahr(sample_excel: Path) -> None:
    participants = parse_excel(sample_excel)
    bob = next(p for p in participants if p.user_id == "u2")
    assert bob.results[0].passed is True


def test_parse_achieved_points_as_string(tmp_path: Path) -> None:
    """achievedPoints stored as a text string must be parsed correctly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(
        [
            "userId",
            "displayName",
            "email",
            "matriculationNumber",
            "groupId",
            "groupName",
            "hasAdmission",
            "rule",
            "assignmentType",
            "achievedPoints",
            "passed",
        ]
    )
    ws.append(
        [
            "u9",
            "Test",
            "t@test.com",
            "999999",
            "g9",
            "G9",
            "WAHR",
            "INDIVIDUAL_PERCENT_WITH_ALLOWED_FAILURES",
            "HOMEWORK",
            "3",  # string, not int
            "FALSCH",
        ]
    )
    path = tmp_path / "string_points.xlsx"
    wb.save(path)

    participants = parse_excel(path)
    assert participants[0].results[0].achieved_points == 3


def test_missing_column_raises(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["userId", "displayName"])  # incomplete headers
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="missing required columns"):
        parse_excel(path)


def test_empty_workbook_returns_empty(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    # Write only headers, no data rows
    ws.append(
        [
            "userId",
            "displayName",
            "email",
            "matriculationNumber",
            "groupId",
            "groupName",
            "hasAdmission",
            "rule",
            "assignmentType",
            "achievedPoints",
            "passed",
        ]
    )
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    assert parse_excel(path) == []


def test_multiple_results_per_participant(tmp_path: Path) -> None:
    """A participant with two result rows gets both Results in their tuple."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(
        [
            "userId",
            "displayName",
            "email",
            "matriculationNumber",
            "groupId",
            "groupName",
            "hasAdmission",
            "rule",
            "assignmentType",
            "achievedPoints",
            "passed",
        ]
    )
    for rule in ("INDIVIDUAL_PERCENT_WITH_ALLOWED_FAILURES", "OTHER_RULE"):
        ws.append(
            [
                "u1",
                "Alice",
                "a@uni.de",
                "111",
                "g1",
                "G1",
                "WAHR",
                rule,
                "HOMEWORK",
                2,
                "FALSCH",
            ]
        )
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    participants = parse_excel(path)
    assert len(participants) == 1
    assert len(participants[0].results) == 2
